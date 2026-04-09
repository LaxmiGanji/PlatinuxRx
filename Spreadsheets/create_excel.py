import sys
import subprocess
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def create_spreadsheet():
    wb = Workbook()
    
    # Sheet 1: ticket
    ws_ticket = wb.active
    ws_ticket.title = "ticket"
    
    ticket_headers = ["ticket_id", "created_at", "closed_at", "outlet_id", "cms_id"]
    ws_ticket.append(ticket_headers)
    
    # Insert some sample data (from the assignment screen and some extra cases for same-day/same-hour)
    ticket_data = [
        ["isu-sjd-457", "2021-08-19 16:45:43", "2021-08-22 12:33:32", "wrqy-juv-978", "vew-iuvd-12"],
        ["qer-fal-092", "2021-08-21 11:09:22", "2021-08-21 17:13:45", "8woh-k3u-23b", "vew-iuvd-13"],
        ["abc-def-123", "2021-08-21 14:15:00", "2021-08-21 14:45:00", "wrqy-juv-978", "vew-iuvd-14"]
    ]
    for row in ticket_data:
        ws_ticket.append(row)
        
    # Helper columns for time analysis (Q2)
    # F: Same Day? => Check if INT(created_at) == INT(closed_at)
    # G: Same Hour? => Check if TIME components have the same hour
    ws_ticket["F1"] = "Same Day?"
    ws_ticket["G1"] = "Same Hour?"
    
    for row in range(2, len(ticket_data) + 2):
        ws_ticket[f"F{row}"] = f"=INT(B{row})=INT(C{row})"
        ws_ticket[f"G{row}"] = f"=AND(F{row}, HOUR(B{row})=HOUR(C{row}))"
        
    # Sheet 2: feedbacks
    ws_feedbacks = wb.create_sheet("feedbacks")
    feedback_headers = ["cms_id", "feedback_at", "feedback_rating", "ticket_created_at"]
    ws_feedbacks.append(feedback_headers)
    
    feedback_data = [
        ["vew-iuvd-12", "2021-08-21 13:26:48", 3],
        ["vew-iuvd-13", "2021-08-22 10:00:00", 4],
        ["vew-iuvd-14", "2021-08-21 15:00:00", 5]
    ]
    
    for row_idx, row in enumerate(feedback_data, start=2):
        ws_feedbacks.append(row)
        # VLOOKUP for ticket_created_at (Q1) -> using standard INDEX-MATCH for right-to-left lookup compatibility.
        ws_feedbacks[f"D{row_idx}"] = f'=INDEX(ticket!B:B, MATCH(A{row_idx}, ticket!E:E, 0))'

    # Auto-adjust column widths for better readability
    for sheet in wb.worksheets:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, "Ticket_Analysis.xlsx")
    wb.save(file_path)
    print(f"Spreadsheet generated successfully at:\n{file_path}")

if __name__ == "__main__":
    create_spreadsheet()
